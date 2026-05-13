#!/usr/bin/env python3
"""Batch download reconstruct audio URLs from an Excel column.

Default target:
    Q column -> algo_audio_reconstruct_event_result.output_url

Example:
    python3 download_reconstruct_audio.py "/Users/xy/Downloads/4月20到5月9日反馈数据.xlsx"
    python3 download_reconstruct_audio.py "/Users/xy/Downloads/4月20到5月9日反馈数据.xlsx" --dry-run
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import time
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urlparse
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string


DEFAULT_KEY = "algo_audio_reconstruct_event_result"
DEFAULT_URL_FIELD = "output_url"
URL_RE = re.compile(r"https?://[^\s\"']+")


def load_sheet(path: Path, sheet_name: str | None):
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    # Some exported xlsx files have a wrong dimension like A1 even when the
    # sheet contains real data. This forces openpyxl to scan all cells.
    worksheet.reset_dimensions()
    worksheet.calculate_dimension(force=True)
    return workbook, worksheet


def regex_extract_url(raw: str, key: str, url_field: str) -> str | None:
    block_pattern = re.compile(
        rf'"{re.escape(key)}"\s*:\s*\{{.*?"{re.escape(url_field)}"\s*:\s*"([^"\\]*(?:\\.[^"\\]*)*)"',
        re.S,
    )
    match = block_pattern.search(raw)
    if match:
        return match.group(1).replace("\\/", "/")

    key_index = raw.find(key)
    if key_index < 0:
        return None

    tail = raw[key_index:]
    url_match = URL_RE.search(tail)
    if url_match:
        return url_match.group(0).replace("\\/", "/")
    return None


def extract_url(value: Any, key: str, url_field: str) -> str | None:
    if value is None:
        return None

    raw = str(value).strip()
    if not raw:
        return None

    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return regex_extract_url(raw, key, url_field)

    if not isinstance(data, dict):
        return regex_extract_url(raw, key, url_field)

    target = data.get(key)
    if isinstance(target, dict):
        url = target.get(url_field)
        return str(url) if url else None
    return None


def safe_filename_from_url(url: str, row_number: int) -> str:
    parsed = urlparse(url)
    name = Path(unquote(parsed.path)).name or f"audio_{row_number}.wav"
    stem = Path(name).stem
    suffix = Path(name).suffix or ".wav"
    safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "_", stem).strip("._-")
    if not safe_stem:
        safe_stem = "audio"
    return f"row_{row_number:04d}_{safe_stem}{suffix}"


def iter_urls(
    xlsx_path: Path,
    sheet_name: str | None,
    column: str,
    start_row: int,
    key: str,
    url_field: str,
):
    workbook, worksheet = load_sheet(xlsx_path, sheet_name)
    try:
        column_index = column_index_from_string(column)
        for row in worksheet.iter_rows(min_row=start_row, min_col=column_index, max_col=column_index):
            cell = row[0]
            url = extract_url(cell.value, key, url_field)
            if url:
                yield cell.row, url
    finally:
        workbook.close()


def download_one(url: str, output_path: Path, timeout: float) -> int:
    request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(request, timeout=timeout) as response:
        data = response.read()
    output_path.write_bytes(data)
    return len(data)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Download audio URLs from algo_audio_reconstruct_event_result in an xlsx column."
    )
    parser.add_argument("xlsx_path", type=Path, help="Excel file path.")
    parser.add_argument("--sheet", default=None, help="Sheet name. Defaults to active sheet.")
    parser.add_argument("--column", default="Q", help="Column to parse. Default: Q.")
    parser.add_argument("--start-row", type=int, default=2, help="First data row. Default: 2.")
    parser.add_argument("--key", default=DEFAULT_KEY, help=f"JSON object key. Default: {DEFAULT_KEY}.")
    parser.add_argument(
        "--url-field",
        default=DEFAULT_URL_FIELD,
        help=f"URL field inside key. Default: {DEFAULT_URL_FIELD}.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("downloads/reconstruct_audio"),
        help="Directory for downloaded audio. Default: downloads/reconstruct_audio.",
    )
    parser.add_argument("--manifest", type=Path, default=None, help="CSV manifest path.")
    parser.add_argument("--dry-run", action="store_true", help="Only list URLs, do not download.")
    parser.add_argument("--overwrite", action="store_true", help="Re-download existing files.")
    parser.add_argument("--limit", type=int, default=None, help="Only process the first N URLs.")
    parser.add_argument("--timeout", type=float, default=60.0, help="Download timeout seconds.")
    parser.add_argument("--retries", type=int, default=2, help="Retries per URL after the first try.")
    parser.add_argument("--sleep", type=float, default=0.2, help="Sleep seconds between downloads.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    xlsx_path = args.xlsx_path.expanduser()
    output_dir = args.output_dir.expanduser()
    manifest_path = (
        args.manifest.expanduser()
        if args.manifest
        else output_dir / "download_manifest.csv"
    )

    entries = list(
        iter_urls(
            xlsx_path=xlsx_path,
            sheet_name=args.sheet,
            column=args.column,
            start_row=args.start_row,
            key=args.key,
            url_field=args.url_field,
        )
    )
    if args.limit is not None:
        entries = entries[: args.limit]

    print(f"Found {len(entries)} URL(s).")
    if args.dry_run:
        for row_number, url in entries:
            print(f"row {row_number}: {url}")
        return

    output_dir.mkdir(parents=True, exist_ok=True)
    manifest_path.parent.mkdir(parents=True, exist_ok=True)

    rows: list[dict[str, Any]] = []
    ok = 0
    skipped = 0
    failed = 0

    for index, (row_number, url) in enumerate(entries, start=1):
        filename = safe_filename_from_url(url, row_number)
        output_path = output_dir / filename
        record: dict[str, Any] = {
            "index": index,
            "row": row_number,
            "url": url,
            "filename": filename,
            "path": str(output_path),
            "status": "",
            "bytes": "",
            "error": "",
        }

        if output_path.exists() and not args.overwrite:
            record["status"] = "skipped_exists"
            skipped += 1
            print(f"[{index}/{len(entries)}] skip row {row_number}: {filename}")
            rows.append(record)
            continue

        last_error = ""
        for attempt in range(args.retries + 1):
            try:
                byte_count = download_one(url, output_path, timeout=args.timeout)
                record["status"] = "ok"
                record["bytes"] = byte_count
                ok += 1
                print(f"[{index}/{len(entries)}] ok row {row_number}: {filename}")
                break
            except Exception as exc:  # noqa: BLE001 - keep manifest error readable.
                last_error = str(exc)
                if attempt < args.retries:
                    time.sleep(args.sleep)
        else:
            record["status"] = "failed"
            record["error"] = last_error
            failed += 1
            print(f"[{index}/{len(entries)}] failed row {row_number}: {last_error}")

        rows.append(record)
        time.sleep(args.sleep)

    with manifest_path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=["index", "row", "url", "filename", "path", "status", "bytes", "error"],
        )
        writer.writeheader()
        writer.writerows(rows)

    print(
        f"Done. ok={ok}, skipped={skipped}, failed={failed}, "
        f"manifest={manifest_path}"
    )


if __name__ == "__main__":
    main()
